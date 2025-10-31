from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Google Sheets API Setup
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
SHEET_ID = '11Gj6Z_U2_uj9SuPSkQuzx1g3Nc80QLYFzynzoaXI1jo'
RANGE_NAME = 'Form Responses 1!A:D'  # Adjust based on your columns

# Load credentials
creds = Credentials.from_service_account_file('service_account.json', scopes=SCOPES)
service = build('sheets', 'v4', credentials=creds)

# Fetch data
sheet = service.spreadsheets()
result = sheet.values().get(spreadsheetId=SHEET_ID, range=RANGE_NAME).execute()
values = result.get('values', [])

if not values or len(values) < 2:
    print("⚠️ No responses found in Google Form.")
else:
    df = pd.DataFrame(values[1:], columns=values[0])
    df.to_excel('form_data.xlsx', index=False)
    print("✅ Google Form data saved to form_data.xlsx")

    # Define your required skills
    required_skills = {"Python", "Machine Learning", "Data Analysis", "JavaScript"}

    # Load workbook for coloring
    wb = load_workbook("form_data.xlsx")
    ws = wb.active

    # Define colors
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Find the "Skills" column
    skill_col = None
    for idx, col_name in enumerate(df.columns):
        if "skill" in col_name.lower():
            skill_col = idx + 1  # +1 because Excel is 1-indexed
            break

    if skill_col:
        # Highlight cells based on skill match
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=skill_col)
            skills_list = {s.strip().title() for s in cell.value.split(",")} if cell.value else set()

            missing_skills = required_skills - skills_list
            extra_skills = skills_list - required_skills

            if missing_skills:
                cell.fill = red_fill
                cell.value += f"  ❌ Missing: {', '.join(missing_skills)}"
            elif extra_skills:
                cell.fill = green_fill
                cell.value += f"  ✅ Extra: {', '.join(extra_skills)}"

    wb.save("form_data_flagged.xlsx")
    print("📊 Skills checked and saved to form_data_flagged.xlsx")
